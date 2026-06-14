"""
Water Bill — Previous-Month Per-Litre Cost Calculator
=====================================================

Streamlit app for a residential society.

Flow:
  1. Upload the monthly Consumption Report (.xlsx)  -- filename must carry the month & year,
     e.g. "Consumption Report May-2026.xlsx".
  2. Upload the Reimbursement template (.csv).
  3. Enter the three water-supply bills for the previous month:
        - Cauvery water bill
        - SSE water tanker bill
        - CKE water tanker bill
  4. App validates that the consumption sheet is for the PREVIOUS month
     (derived from today's date). Current / future / older months are rejected.
  5. Per-litre cost = (Cauvery + SSE + CKE) / grand total of the updated reimbursement sheet,
     where each flat's reading is taken from the consumption sheet's Total, merged flats are
     consolidated, and any flat whose total is below 50 L is billed as 0.

The fill logic mirrors the manually-verified sheet:
  - flat numbers normalised  "A 001" <-> "A-001"
  - merged meters            "C 1303 & 1403" -> CSV row "C-1303"
  - sub-threshold flats       total < 50 L  ->  0
"""

import calendar
import csv
import io
import re
from datetime import date

import openpyxl

# --------------------------------------------------------------------------- #
# Pure logic (importable / testable without Streamlit)
# --------------------------------------------------------------------------- #

SUB_THRESHOLD_LITRES = 50
_MONTH_BY_NAME = {name.lower(): i for i, name in enumerate(calendar.month_name) if name}
_MONTH_BY_NAME.update({abbr.lower(): i for i, abbr in enumerate(calendar.month_abbr) if abbr})


def previous_month(today: date) -> tuple[int, int]:
    """Return (year, month) of the month immediately before `today`."""
    year, month = today.year, today.month - 1
    if month == 0:
        year, month = year - 1, 12
    return year, month


def parse_month_from_filename(filename: str) -> tuple[int, int] | None:
    """Extract (year, month) from a filename like 'Consumption Report May-2026.xlsx'.

    Returns None if a month name or a 4-digit year cannot be found.
    """
    stem = re.sub(r"\.[^.]+$", "", filename)          # drop extension
    tokens = re.findall(r"[A-Za-z]+|\d{4}", stem)

    month = next((_MONTH_BY_NAME[t.lower()] for t in tokens if t.lower() in _MONTH_BY_NAME), None)
    year = next((int(t) for t in tokens if t.isdigit() and len(t) == 4), None)
    if month is None or year is None:
        return None
    return year, month


def _normalise_flat(value: str) -> str:
    """'A 001' -> 'A-001';  'C 1303 & 1403' -> 'C-1303' (merged meter -> first flat)."""
    first = re.split(r"&", str(value))[0]
    return re.sub(r"\s+", "-", first.strip()).upper()


def load_consumption_totals(file) -> dict[str, float]:
    """Read the consumption workbook and return {normalised_flat: total_litres}.

    `file` may be a path or a file-like object. The grand-total row is skipped.
    Apartment and Total columns are located by header name for robustness.
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb["data"] if "data" in wb.sheetnames else wb.active

    header = [c.value for c in ws[1]]
    try:
        apt_col = header.index("Apartment") + 1
    except ValueError:
        apt_col = 1
    try:
        total_col = header.index("Total") + 1
    except ValueError:
        total_col = ws.max_column

    totals: dict[str, float] = {}
    for row in range(2, ws.max_row + 1):
        apt = ws.cell(row=row, column=apt_col).value
        if apt in (None, ""):
            continue
        apt = str(apt).strip()
        if apt.lower() == "total":               # grand-total row at the bottom
            continue
        raw = ws.cell(row=row, column=total_col).value
        totals[_normalise_flat(apt)] = float(raw) if raw is not None else 0.0
    return totals


def billable_reading(raw_total: float) -> float:
    """Apply the sub-50-litre rule."""
    if raw_total is None or raw_total < SUB_THRESHOLD_LITRES:
        return 0.0
    return float(raw_total)


def build_updated_sheet(csv_text: str, totals: dict[str, float], description: str) -> dict:
    """Fill the reimbursement sheet and sum its Current Reading column.

    For each data row: matched flats get the billable consumption total in 'Current Reading';
    rows with no consumption match (clubhouse / non-member / common-area) keep their existing
    reading. Every data row's 'Description' is set to `description`.

    Returns a breakdown dict including `csv_text` — the updated sheet ready for download.
    """
    rows = list(csv.reader(io.StringIO(csv_text)))
    CUR, DESC = 2, 4                               # 'Current Reading' / 'Description' columns

    grand_total = 0.0
    matched = zeroed = unmatched = 0
    unmatched_rows: list[str] = []

    for row in rows[2:]:                           # first two lines are header / section labels
        if not row or not row[0].strip():
            continue
        house = row[0].strip()
        key = _normalise_flat(house)
        if key in totals:
            reading = billable_reading(totals[key])
            row[CUR] = f"{reading:.3f}"            # write reading, keep the 3-decimal column style
            if reading == 0.0:
                zeroed += 1
            matched += 1
        else:
            try:
                reading = float(row[CUR]) if len(row) > CUR and row[CUR].strip() else 0.0
            except ValueError:
                reading = 0.0
            unmatched += 1
            unmatched_rows.append(house)           # reading left unchanged
        if len(row) > DESC:
            row[DESC] = description                # update description for every data row
        grand_total += reading

    out = io.StringIO()
    csv.writer(out).writerows(rows)

    return {
        "grand_total": grand_total,
        "matched": matched,
        "zeroed": zeroed,
        "unmatched": unmatched,
        "unmatched_rows": unmatched_rows,
        "csv_text": out.getvalue(),
    }


# --------------------------------------------------------------------------- #
# Streamlit UI
# --------------------------------------------------------------------------- #

def main() -> None:
    import streamlit as st

    st.set_page_config(page_title="Water Bill — Per-Litre Cost", page_icon="💧")
    st.title("💧 Water Bill — Per-Litre Cost (Previous Month)")

    today = date.today()
    py, pm = previous_month(today)
    prev_label = f"{calendar.month_name[pm]} {py}"
    cur_label = f"{calendar.month_name[today.month]} {today.year}"

    st.info(
        f"**Billing month:** {prev_label}  \n"
        f"Today is {today:%d %b %Y}. Bills can only be generated for the **previous month**. "
        f"Current-month ({cur_label}) generation is blocked."
    )

    st.subheader("1 · Upload sheets")
    cons_file = st.file_uploader(
        "Consumption Report (.xlsx) — filename must include the month & year, "
        "e.g. 'Consumption Report May-2026.xlsx'",
        type=["xlsx"],
    )
    csv_file = st.file_uploader("Reimbursement template (.csv)", type=["csv"])

    st.subheader("2 · Previous-month water-supply bills (₹)")
    c1, c2, c3 = st.columns(3)
    cauvery = c1.number_input("Cauvery water bill", min_value=0.0, step=100.0, format="%.2f")
    sse = c2.number_input("SSE water tanker bill", min_value=0.0, step=100.0, format="%.2f")
    cke = c3.number_input("CKE water tanker bill", min_value=0.0, step=100.0, format="%.2f")

    st.subheader("3 · Generate")
    if st.button(f"Generate {prev_label} bill", type="primary"):
        # Compute on click and stash in session_state so the result (and its download
        # button) survive the rerun that Streamlit triggers when the button is pressed.
        st.session_state["result"] = _generate(
            cons_file, csv_file, cauvery, sse, cke, today, (py, pm), prev_label
        )

    result = st.session_state.get("result")
    if result is None:
        return
    if result.get("error"):
        st.error(result["error"])
        return

    _render_result(st, result)


def _generate(cons_file, csv_file, cauvery, sse, cke, today, prev, prev_label) -> dict:
    """Validate inputs and compute the result. Returns {'error': msg} or a result dict."""
    py, pm = prev
    if cons_file is None or csv_file is None:
        return {"error": "Please upload both the consumption report and the reimbursement template."}

    parsed = parse_month_from_filename(cons_file.name)
    if parsed is None:
        return {"error": (
            f"Could not read a month and year from the filename '{cons_file.name}'. "
            "Rename it to include both, e.g. 'Consumption Report May-2026.xlsx'."
        )}

    file_year, file_month = parsed
    file_label = f"{calendar.month_name[file_month]} {file_year}"
    if (file_year, file_month) == (today.year, today.month):
        return {"error": (
            f"❌ The uploaded sheet is for the **current month** ({file_label}). "
            f"Current-month bills cannot be generated — upload the previous month ({prev_label})."
        )}
    if (file_year, file_month) != (py, pm):
        return {"error": (
            f"❌ The uploaded sheet is for **{file_label}**, but only the previous month "
            f"(**{prev_label}**) can be billed. Please upload the correct sheet."
        )}

    total_cost = cauvery + sse + cke
    if total_cost <= 0:
        return {"error": "Enter at least one non-zero water-supply bill amount."}

    try:
        totals = load_consumption_totals(cons_file)
        csv_text = csv_file.getvalue().decode("utf-8-sig")
        breakdown = build_updated_sheet(csv_text, totals, prev_label)
    except Exception as exc:  # noqa: BLE001 — surface any parse error to the user
        return {"error": f"Failed to process the sheets: {exc}"}

    grand_total = breakdown["grand_total"]
    if grand_total <= 0:
        return {"error": "Total billable consumption is 0 — cannot compute a per-litre cost."}

    return {
        "prev_label": prev_label,
        "cauvery": cauvery, "sse": sse, "cke": cke,
        "total_cost": total_cost,
        "grand_total": grand_total,
        "per_litre": total_cost / grand_total,
        "breakdown": breakdown,
        "download_name": f"Reimbursement Towards Monthly Water Bill - "
                         f"{calendar.month_name[pm]}-{py}.csv",
    }


def _render_result(st, r: dict) -> None:
    b = r["breakdown"]
    st.success(f"Per-litre water cost for **{r['prev_label']}**")
    st.metric("Per-litre cost", f"₹ {r['per_litre']:,.4f} / litre")

    st.markdown("#### Calculation breakdown")
    st.markdown(
        f"- Cauvery bill: ₹ {r['cauvery']:,.2f}\n"
        f"- SSE tanker bill: ₹ {r['sse']:,.2f}\n"
        f"- CKE tanker bill: ₹ {r['cke']:,.2f}\n"
        f"- **Total supply cost:** ₹ {r['total_cost']:,.2f}\n"
        f"- **Total billable consumption (updated sheet grand total):** "
        f"{r['grand_total']:,.0f} litres\n"
        f"- **Per-litre cost = {r['total_cost']:,.2f} ÷ {r['grand_total']:,.0f} = "
        f"₹ {r['per_litre']:,.6f} / litre**"
    )

    st.markdown("#### Download")
    st.download_button(
        "⬇️ Download updated reimbursement sheet (.csv)",
        data=b["csv_text"].encode("utf-8-sig"),
        file_name=r["download_name"],
        mime="text/csv",
        type="primary",
    )

    with st.expander("Sheet processing details"):
        st.write(
            f"- Flats matched from consumption sheet: **{b['matched']}**\n"
            f"- Flats below {SUB_THRESHOLD_LITRES} L billed as 0: **{b['zeroed']}**\n"
            f"- Non-flat rows kept as-is (clubhouse / non-member / common area): "
            f"**{b['unmatched']}**"
        )
        if b["unmatched_rows"]:
            st.caption("Non-flat / unmatched rows: " + ", ".join(b["unmatched_rows"]))


if __name__ == "__main__":
    main()
